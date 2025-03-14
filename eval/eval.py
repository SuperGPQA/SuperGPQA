import json
import re
import argparse
import os
from prettytable import PrettyTable
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from tqdm import tqdm
import timeout_decorator
import multiprocessing
import time
from functools import partial

@timeout_decorator.timeout(5)  # 5 seconds timeout
def safe_regex_search(pattern, text, flags=0):
    """
    TODO: The optimal solution for timeout detection is to use the 'regex' library instead of 're' for regular expression matching.
    However, since the 'regex' and 're' libraries handle regex parsing differently, it has not been adopted for now.
    
    Issue: The current implementation using 'timeout_decorator' does not work on Windows platforms.
    Reason: 'timeout_decorator' relies on signal-based timeouts, which are only supported on Unix-based systems and do not work on Windows.
    """
    try:
        return re.search(pattern, text, flags)
    except timeout_decorator.TimeoutError:
        print(f"Regex match timeout: pattern={pattern}, text={text[:100]}...")
        return None
    except Exception as e:
        print(f"Regex match error: {str(e)}")
        return None

def extract_option_labels(text, options='ABCDEFGHIJ'):
    if not isinstance(text, str) or not isinstance(options, str):
        return 'error'
    
    text = text.rstrip()
    last_line = text.split('\n')[-1]
    
    option_str = ''.join([chr(65 + i) for i in range(len(options))]) if options else 'ABCDEFGHIJ'
    
    patterns = [
        # e.g. "The final answer to this question is: A."
        #      "The best option is $\boxed{B}:"
        #      "The correct answer is (C)."
        f'[Tt]he\s+(?:\w+\s+)?(?:answer|option)(?:\w+\s+)?\s+is?:?\s*(?:[\*\$\\{{(\[\\\\(]*?(?:(?:\\\\boxed|\\\\mathbf|\\\\mathrm|\\\\text){{)?)*\s*([{option_str}])(?:\\\\?\}}?\$?\)?\]?\}}?)*(?:[\s:\.\*)]|$)',

        # e.g. "ANSWER: A"
        #      "Answer: $\boxed{B}."
        #      "ANSWER: (C):"
        f'(?i:Answer)[\*\s]*:\s*(?:[\*\$\\{{(\[\\\\(]*?(?:(?:\\\\boxed|\\\\mathbf|\\\\mathrm|\\\\text){{)?)*\s*([{option_str}])(?:\\\\?\}}?\$?\)?\]?\}}?)*(?:[\s:\.\*)]|$)',
        
        # e.g. "A"
        #      "$\boxed{B}$"
        #      "(C)."
        #      "[D]:"
        f'^[^\w\r\n]*(?:[\*\$\\{{(\[\\\\(]*?(?:(?:\\\\boxed|\\\\mathbf|\\\\mathrm|\\\\text){{)?)*\s*([{option_str}])(?:\\\\?\}}?\$?\)?\]?\}}?)*(?:[\s:\.\*)]|$)',
    ]

    for pattern in patterns:
        match = safe_regex_search(pattern, last_line, re.IGNORECASE)
        if match:
            return match.group(1)
    
    for pattern in patterns:
        match = safe_regex_search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)
    
    return None

def extract_option_content(text, options_content=None):
    if not isinstance(text, str) or not isinstance(options_content, list):
        return 'error'
    
    escaped_options_content = [re.escape(option_content) for option_content in options_content]
    escaped_options_content_str = '|'.join(escaped_options_content)
    
    text = text.rstrip()
    last_line = text.split('\n')[-1]
        
    patterns = [
        f'[Tt]he\s+(?:\w+\s+)?(?:answer|option)(?:\w+\s+)?\s+is:?\s*(?:[\*\$\\{{\(\[\\\\(]*?(?:(?:\\\\boxed|\\\\mathbf|\\\\mathrm|\\\\text){{)?)*\s*({escaped_options_content_str})(?:\\\\?\}}?\$?\)?\]?\}}?)*(?:[\s:\.\*)]|$)',
        
        f'(?i:Answer)\s*(?:[\*\$\\{{\(\[\\\\(]*?(?:(?:\\\\boxed|\\\\mathbf|\\\\mathrm|\\\\text){{)?)*\s*({escaped_options_content_str})(?:\\\\?\}}?\$?\)?\]?\}}?)*(?:[\s:\.\*)]|$)',
        
        f'^[^\w\r\n]*(?:[\*\$\\{{\(\[\\\\(]*?(?:(?:\\\\boxed|\\\\mathbf|\\\\mathrm|\\\\text){{)?)*\s*({escaped_options_content_str})(?:\\\\?\}}?\$?\)?\]?\}}?)*(?:[\s:\.\*)]|$)',
    ]
    
    for pattern in patterns:
        match = safe_regex_search(pattern, last_line)
        if match:
            if match.group(1) in escaped_options_content:
                return options_content[escaped_options_content.index(match.group(1))]
            else:
                return match.group(1)
    
    for pattern in patterns:
        match = safe_regex_search(pattern, text)
        if match:
            if match.group(1) in escaped_options_content:
                return options_content[escaped_options_content.index(match.group(1))]
            else:
                return match.group(1)
    
    return None

def calculate_accuracy(file_path, save_dir, mode):
    data = []
    acc = 0
    count = 0
    err = 0
    miss = 0
    acc_difficulty = {"hard": 0, "middle": 0, "easy": 0}
    count_difficulty = {"hard": 0, "middle": 0, "easy": 0}
    
    stats = {
        'discipline': {},
        'field': {},
        'subfield': {}
    }
    
    with open(file_path, "r") as file:
        for line in tqdm(file, desc=f"Reading {os.path.basename(file_path)} data", leave=False):
            data.append(json.loads(line))
    
    if not data:
        print(f"Warning: No data found in {file_path}")
        return 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, stats
    
    for sample in tqdm(data, desc=f"Processing {os.path.basename(file_path)} samples", leave=False):
        if mode == 'zero-shot':
            predict = extract_option_labels(sample["response"], 'ABCDEFGHIJ')
            if predict == None:
                predict = extract_option_content(sample["response"], sample["options"])
                predict = chr(sample["options"].index(predict) + 65) if predict else None
            sample["extracted_answer"] = predict
        elif mode == 'five-shot':
            response = sample["response"].split('Question:')[0]
            predict = extract_option_labels(response, 'ABCDEFGHIJ')
            if predict == None:
                predict = extract_option_content(response, sample["options"])
                predict = chr(sample["options"].index(predict) + 65) if predict else None
            if predict == None:
                predict = extract_option_labels(sample["response"], 'ABCDEFGHIJ')
                if predict == None:
                    predict = extract_option_content(sample["response"], sample["options"])
                    predict = chr(sample["options"].index(predict) + 65) if predict else None
            sample["extracted_answer"] = predict
        
        discipline = sample.get("discipline", "unknown")
        field = sample.get("field", "unknown")
        subfield = sample.get("subfield", "unknown")
        difficulty = sample.get("difficulty", "unknown")
        
        for level, key in [
            ('discipline', discipline),
            ('field', f"{discipline}/{field}"),
            ('subfield', f"{discipline}/{field}/{subfield}")
        ]:
            if key not in stats[level]:
                stats[level][key] = {
                    "correct": 0, 
                    "total": 0, 
                    "miss": 0, 
                    "error": 0,
                    "discipline": discipline,
                    "field": field,
                    "subfield": subfield,
                    "difficulty": {
                        "easy": {"correct": 0, "total": 0},
                        "middle": {"correct": 0, "total": 0},
                        "hard": {"correct": 0, "total": 0}
                    }
                }
            
            stats[level][key]["total"] += 1
            stats[level][key]["difficulty"][difficulty]["total"] += 1
            
            answer_letter = sample["answer_letter"]
            
            if predict and answer_letter == predict:
                acc += 1
                acc_difficulty[difficulty] += 1
                sample["status"] = "correct"
                stats[level][key]["correct"] += 1
                stats[level][key]["difficulty"][difficulty]["correct"] += 1
            elif predict == None or predict == "":
                miss += 1
                sample["status"] = "miss"
                stats[level][key]["miss"] += 1
            elif predict == 'error':
                err += 1
                sample["status"] = "error"
                stats[level][key]["error"] += 1
            else:
                sample["status"] = "incorrect"
            count += 1
            count_difficulty[difficulty] += 1
    
    if count == 0:
        print(f"Warning: No valid samples found in {file_path}")
        return 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, stats
    
    accuracy = acc / count
    error_rate = err / count
    miss_rate = miss / count
    hard_accuracy = acc_difficulty["hard"] / count_difficulty["hard"]
    middle_accuracy = acc_difficulty["middle"] / count_difficulty["middle"]
    easy_accuracy = acc_difficulty["easy"] / count_difficulty["easy"]
    
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, os.path.basename(file_path))
    with open(save_path, "w") as file:
        for sample in data:
            json.dump(sample, file)
            file.write("\n")
    
    return accuracy, error_rate, miss_rate, hard_accuracy, middle_accuracy, easy_accuracy, stats

def calculate_total_row(hierarchy_stats, model_results, metric_name):
    """Calculate overall summary row, including sample-wise and weighted average across dimensions"""
    total_rows = []
    
    # Calculate total counts across dimensions
    total_samples = 0
    if metric_name in ['Hard', 'Middle', 'Easy']:
        total_subfields = sum(subfield['difficulty'][metric_name.lower()]['total'] > 0 for subfield in hierarchy_stats['subfield'].values())
        total_fields = sum(field['difficulty'][metric_name.lower()]['total'] > 0 for field in hierarchy_stats['field'].values())
        total_disciplines = sum(discipline['difficulty'][metric_name.lower()]['total'] > 0 for discipline in hierarchy_stats['discipline'].values())
    else:
        total_subfields = len(hierarchy_stats['subfield'])
        total_fields = len(hierarchy_stats['field'])
        total_disciplines = len(hierarchy_stats['discipline'])
    
    # Calculate total sample count
    for discipline_stats in hierarchy_stats['discipline'].values():
        if metric_name in ['Hard', 'Middle', 'Easy']:
            total_samples += discipline_stats['difficulty'][metric_name.lower()]['total']
        else:
            total_samples += discipline_stats['total']
    
    if metric_name == 'Accuracy':
        row_types = [
            (f'Overall (sample-wise) (Total samples: {total_samples})', 'sample'),
            (f'Overall (subfield-wise) (Total subfields: {total_subfields})', 'subfield'),
            (f'Overall (field-wise) (Total fields: {total_fields})', 'field'),
            (f'Overall (discipline-wise) (Total disciplines: {total_disciplines})', 'discipline')
        ]
    elif metric_name in ['Hard', 'Middle', 'Easy']:
        row_types = [
            (f'Overall (sample-wise) (Total {metric_name.lower()} samples: {total_samples})', 'sample'),
            (f'Overall (subfield-wise) (Total {metric_name.lower()} subfields: {total_subfields})', 'subfield'),
            (f'Overall (field-wise) (Total {metric_name.lower()} fields: {total_fields})', 'field'),
            (f'Overall (discipline-wise) (Total {metric_name.lower()} disciplines: {total_disciplines})', 'discipline')
        ]
    else:  # Error Rate and Miss Rate
        row_types = [(f'Overall (Total samples: {total_samples})', 'sample')]
    
    for row_name, stat_type in row_types:
        total_row = {
            'Discipline': row_name,
            'Field': '',
            'Subfield': ''
        }
        
        for model_name in model_results.keys():
            for mode in model_results[model_name].keys():
                if stat_type == 'sample':
                    # sample-wise statistics (weighted by sample count)
                    stats = {'total': 0, 'correct': 0, 'error': 0, 'miss': 0}
                    
                    for discipline_stats in hierarchy_stats['discipline'].values():
                        if 'model_stats' in discipline_stats and model_name in discipline_stats['model_stats']:
                            curr_stats = discipline_stats['model_stats'][model_name].get(mode, {})
                            
                            if metric_name in ['Hard', 'Middle', 'Easy']:
                                difficulty_stats = curr_stats.get('difficulty', {}).get(metric_name.lower(), {})
                                stats['total'] += difficulty_stats.get('total', 0)
                                stats['correct'] += difficulty_stats.get('correct', 0)
                            else:
                                for key in ['total', 'correct', 'error', 'miss']:
                                    stats[key] += curr_stats.get(key, 0)
                    
                    if stats['total'] > 0:
                        if metric_name in ['Hard', 'Middle', 'Easy'] or metric_name == 'Accuracy':
                            value = stats['correct'] / stats['total']
                        elif metric_name == 'Error Rate':
                            value = stats['error'] / stats['total']
                        else:  # Miss Rate
                            value = stats['miss'] / stats['total']
                    else:
                        value = 0
                
                else:
                    # Other dimension statistics (direct average of correct rates across categories)
                    scores = []
                    
                    if stat_type == 'discipline':
                        categories = hierarchy_stats['discipline']
                    elif stat_type == 'field':
                        categories = hierarchy_stats['field']
                    else:  # subfield
                        categories = hierarchy_stats['subfield']
                    
                    for cat_stats in categories.values():
                        if 'model_stats' in cat_stats and model_name in cat_stats['model_stats']:
                            curr_stats = cat_stats['model_stats'][model_name].get(mode, {})
                            
                            if metric_name in ['Hard', 'Middle', 'Easy']:
                                difficulty_stats = curr_stats.get('difficulty', {}).get(metric_name.lower(), {})
                                if difficulty_stats.get('total', 0) > 0:
                                    score = difficulty_stats['correct'] / difficulty_stats['total']
                                    scores.append(score)
                            else:
                                if curr_stats.get('total', 0) > 0:
                                    if metric_name == 'Accuracy':
                                        score = curr_stats['correct'] / curr_stats['total']
                                        scores.append(score)
                    value = sum(scores) / len(scores) if scores else 0
                    
                total_row[f'{model_name}_{mode}'] = f"{value:.2%}"
        
        total_rows.append(total_row)
    
    return total_rows

def create_excel_report_from_stats(model_results, hierarchy_stats, save_path):
    print("Starting Excel report generation...")
    
    # Create six different DataFrames for storing different metrics and difficulties
    metrics = {
        'Accuracy': {'rows': [], 'color': '000000'},  # black
        'Error Rate': {'rows': [], 'color': '000000'},  # black
        'Miss Rate': {'rows': [], 'color': '000000'},   # black
        'Hard': {'rows': [], 'color': '000000'},    # black
        'Middle': {'rows': [], 'color': '000000'},  # black
        'Easy': {'rows': [], 'color': '000000'}     # black
    }
    
    # Organize data by hierarchy
    for discipline in tqdm(sorted(hierarchy_stats['discipline'].keys()), desc="Processing discipline level"):
        discipline_stats = hierarchy_stats['discipline'][discipline]
        discipline_total = discipline_stats['total']
        
        # Get all fields under this discipline
        categories = [k for k in hierarchy_stats['field'].keys() 
                     if k.startswith(f"{discipline}/")]
        
        for field_key in sorted(categories):
            field_stats = hierarchy_stats['field'][field_key]
            field = field_stats['field']
            field_total = field_stats['total']
            
            # Get all subfields under this field
            subcategories = [k for k in hierarchy_stats['subfield'].keys() 
                           if k.startswith(f"{discipline}/{field}/")]
            
            # Add subfield row
            for subfield_key in sorted(subcategories):
                subfield_stats = hierarchy_stats['subfield'][subfield_key]
                
                # Create base row data for each metric
                for metric_name in metrics:
                    if metric_name in ['Hard', 'Middle', 'Easy']:
                        base_row = {
                            'Discipline': discipline,
                            'Field': field,
                            'Subfield': f"{subfield_stats['subfield']} ({subfield_stats['difficulty'][metric_name.lower()]['total']})"
                        }
                    else:
                        base_row = {
                            'Discipline': discipline,
                            'Field': field,
                            'Subfield': f"{subfield_stats['subfield']} ({subfield_stats['total']})"
                        }
                    
                    row_data = base_row.copy()
                    
                    # Add score for each model
                    for model_name in model_results.keys():
                        for mode in model_results[model_name].keys():
                            stats = subfield_stats['model_stats'].get(model_name, {}).get(mode, {})
                            
                            if metric_name in ['Hard', 'Middle', 'Easy']:
                                difficulty_stats = stats.get('difficulty', {}).get(metric_name.lower(), {})
                                if difficulty_stats.get('total', 0) > 0:
                                    value = f"{difficulty_stats['correct'] / difficulty_stats['total']:.2%}"
                                else:
                                    value = '0.00%'
                            else:
                                if stats.get('total', 0) > 0:
                                    if metric_name == 'Accuracy':
                                        value = f"{stats['correct'] / stats['total']:.2%}"
                                    elif metric_name == 'Error Rate':
                                        value = f"{stats['error'] / stats['total']:.2%}"
                                    else:  # Miss Rate
                                        value = f"{stats['miss'] / stats['total']:.2%}"
                                else:
                                    value = '0.00%'
                            
                            row_data[f'{model_name}_{mode}'] = value
                    
                    metrics[metric_name]['rows'].append(row_data)
                
            # Add field summary row
            for metric_name in metrics:
                if metric_name in ['Hard', 'Middle', 'Easy']:
                    field_row = {
                        'Discipline': discipline,
                        'Field': f"{field} (Total: {field_stats['difficulty'][metric_name.lower()]['total']})",
                        'Subfield': ''
                    }
                else:
                    field_row = {
                        'Discipline': discipline,
                        'Field': f"{field} (Total: {field_total})",
                        'Subfield': ''
                    }
                
                for model_name in model_results.keys():
                    for mode in model_results[model_name].keys():
                        stats = field_stats['model_stats'].get(model_name, {}).get(mode, {})
                        
                        if metric_name in ['Hard', 'Middle', 'Easy']:
                            difficulty_stats = stats.get('difficulty', {}).get(metric_name.lower(), {})
                            if difficulty_stats.get('total', 0) > 0:
                                value = f"{difficulty_stats['correct'] / difficulty_stats['total']:.2%}"
                            else:
                                value = '0.00%'
                        else:
                            if stats.get('total', 0) > 0:
                                if metric_name == 'Accuracy':
                                    value = f"{stats['correct'] / stats['total']:.2%}"
                                elif metric_name == 'Error Rate':
                                    value = f"{stats['error'] / stats['total']:.2%}"
                                else:  # Miss Rate
                                    value = f"{stats['miss'] / stats['total']:.2%}"
                            else:
                                value = '0.00%'
                        
                        field_row[f'{model_name}_{mode}'] = value
                
                metrics[metric_name]['rows'].append(field_row)

        # Add discipline summary row
        for metric_name in metrics:
            if metric_name in ['Hard', 'Middle', 'Easy']:
                discipline_row = {
                    'Discipline': f"{discipline} (Total: {discipline_stats['difficulty'][metric_name.lower()]['total']})",
                    'Field': '',
                    'Subfield': ''
                }
            else:
                discipline_row = {
                    'Discipline': f"{discipline} (Total: {discipline_total})",
                    'Field': '',
                    'Subfield': ''
                }
            
            for model_name in model_results.keys():
                for mode in model_results[model_name].keys():
                    stats = discipline_stats['model_stats'].get(model_name, {}).get(mode, {})
                    
                    if metric_name in ['Hard', 'Middle', 'Easy']:
                        difficulty_stats = stats.get('difficulty', {}).get(metric_name.lower(), {})
                        if difficulty_stats.get('total', 0) > 0:
                            value = f"{difficulty_stats['correct'] / difficulty_stats['total']:.2%}"
                        else:
                            value = '0.00%'
                    else:
                        if stats.get('total', 0) > 0:
                            if metric_name == 'Accuracy':
                                value = f"{stats['correct'] / stats['total']:.2%}"
                            elif metric_name == 'Error Rate':
                                value = f"{stats['error'] / stats['total']:.2%}"
                            else:  # Miss Rate
                                value = f"{stats['miss'] / stats['total']:.2%}"
                        else:
                            value = '0.00%'
                    
                    discipline_row[f'{model_name}_{mode}'] = value
            
            metrics[metric_name]['rows'].append(discipline_row)

    # Create DataFrames
    dfs = {metric: pd.DataFrame(data['rows']) for metric, data in metrics.items()}
    
    # Add overall summary row to each DataFrame
    for metric_name, df in dfs.items():
        total_rows = calculate_total_row(hierarchy_stats, model_results, metric_name)
        dfs[metric_name] = pd.concat([df, pd.DataFrame(total_rows)], ignore_index=True)

    # Save to Excel, one sheet per metric
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        for metric_name, df in dfs.items():
            df.to_excel(writer, sheet_name=metric_name, index=False)
            format_worksheet(writer.sheets[metric_name], df, metrics[metric_name]['color'])

    print(f"Report generation completed, Excel file saved: {save_path}")

def format_worksheet(worksheet, df, color):
    """Format worksheet"""
    # Set default font
    for row in worksheet.rows:
        for cell in row:
            cell.font = Font(name='Arial', color='000000')  # Use black font uniformly
    
    # Set background color
    discipline_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    field_fill = PatternFill(start_color='FFFFD4', end_color='FFFFD4', fill_type='solid')
    
    # Overall row background color
    sample_wise_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')  # Bright but not bright blue
    subfield_wise_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # Light blue
    field_wise_fill = PatternFill(start_color='E9EEF5', end_color='E9EEF5', fill_type='solid')  # Lighter blue
    discipline_wise_fill = PatternFill(start_color='F2F5F9', end_color='F2F5F9', fill_type='solid')  # Lightest blue
    error_rate_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Red
    miss_rate_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Gray

    # Set column width
    for column in worksheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Merge cells and apply background color
    current_discipline = None
    discipline_start = None
    current_field = None
    field_start = None
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        discipline = row[0].value
        field = row[1].value
        
        # Process discipline (Discipline) merge
        if discipline and "Total:" in str(discipline):
            # If there was an unmerged discipline row before
            if discipline_start and current_discipline:
                worksheet.merge_cells(f'A{discipline_start}:A{row_idx-1}')
            
            # Apply background color to current total row
            for cell in row:
                cell.fill = discipline_fill
            
            # Reset tracking variables
            current_discipline = None
            discipline_start = None
        elif discipline and discipline != current_discipline:
            # If there was an unmerged discipline row before
            if discipline_start and current_discipline:
                worksheet.merge_cells(f'A{discipline_start}:A{row_idx-1}')
            
            current_discipline = discipline
            discipline_start = row_idx
        
        # Process field (Field) merge
        if field and "Total:" in str(field):
            # If there was an unmerged field row before
            if field_start and current_field:
                worksheet.merge_cells(f'B{field_start}:B{row_idx-1}')
            
            # Apply background color to current total row
            for cell in row:
                cell.fill = field_fill
            
            # Reset tracking variables
            current_field = None
            field_start = None
        elif field and field != current_field:
            # If there was an unmerged field row before
            if field_start and current_field:
                worksheet.merge_cells(f'B{field_start}:B{row_idx-1}')
            
            current_field = field
            field_start = row_idx
    
    # Process last unmerged cells
    last_row = worksheet.max_row
    if discipline_start and current_discipline:
        worksheet.merge_cells(f'A{discipline_start}:A{last_row}')
    if field_start and current_field:
        worksheet.merge_cells(f'B{field_start}:B{last_row}')
    
    # Apply special background color to Overall row
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        cell_value = row[0].value
        if cell_value:
            if 'Overall (sample-wise)' in str(cell_value):
                for cell in row:
                    cell.fill = sample_wise_fill
            elif 'Overall (subfield-wise)' in str(cell_value):
                for cell in row:
                    cell.fill = subfield_wise_fill
            elif 'Overall (field-wise)' in str(cell_value):
                for cell in row:
                    cell.fill = field_wise_fill
            elif 'Overall (discipline-wise)' in str(cell_value):
                for cell in row:
                    cell.fill = discipline_wise_fill
            elif worksheet.title == 'Error Rate' and 'Overall' in str(cell_value):
                for cell in row:
                    cell.fill = error_rate_fill
            elif worksheet.title == 'Miss Rate' and 'Overall' in str(cell_value):
                for cell in row:
                    cell.fill = miss_rate_fill

    # Set value format to keep two decimal places
    for row in worksheet.iter_rows(min_row=2):
        for cell in row[3:]:  # Start from 4th column (skip Discipline, Field, Subfield columns)
            if isinstance(cell.value, str) and '%' in cell.value:
                try:
                    value = float(cell.value.strip('%')) / 100
                    cell.value = f"{value:.2%}"
                except ValueError:
                    pass

    # Set all cells to center alignment
    for row in worksheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

def format_cell_value(stats):
    """Format cell content, return string with acc/error/miss"""
    total = stats['total']
    if total == 0:
        return '0%/0%/0%'
    
    acc = stats['correct'] / total
    error = stats['error'] / total
    miss = stats['miss'] / total
    
    return f"{acc:.1%}/{error:.1%}/{miss:.1%}"

def process_single_file(file_name, args):
    """Process a single file and return its data"""
    try:
        if args.split == '':
            model_name, split, mode = file_name.split('_')
        else:
            if args.split in file_name:
                model_name, mode = file_name.split(f'_{args.split}_')
                split = args.split
                mode = mode.replace('.jsonl', '')
                if mode not in args.mode:
                    return None
            else:
                return None
                
        file_path = os.path.join(args.output_dir, file_name)
        
        data = []
        start_time = time.time()
        with open(file_path, "r") as file:
            for line in file:
                data.append(json.loads(line))
        read_time = time.time() - start_time
        
        regex_start_time = time.time()
        acc = 0
        count = 0
        err = 0
        miss = 0
        acc_difficulty = {"hard": 0, "middle": 0, "easy": 0}
        count_difficulty = {"hard": 0, "middle": 0, "easy": 0}
        stats = {
            'discipline': {},
            'field': {},
            'subfield': {}
        }
        
        for sample in data:
            if mode == 'zero-shot':
                predict = extract_option_labels(sample["response"], 'ABCDEFGHIJ')
                if predict == None:
                    predict = extract_option_content(sample["response"], sample["options"])
                    predict = chr(sample["options"].index(predict) + 65) if predict else None
                sample["extracted_answer"] = predict
            elif mode == 'five-shot':
                response = sample["response"].split('Question:')[0]
                predict = extract_option_labels(response, 'ABCDEFGHIJ')
                if predict == None:
                    predict = extract_option_content(response, sample["options"])
                    predict = chr(sample["options"].index(predict) + 65) if predict else None
                if predict == None:
                    predict = extract_option_labels(sample["response"], 'ABCDEFGHIJ')
                    if predict == None:
                        predict = extract_option_content(sample["response"], sample["options"])
                        predict = chr(sample["options"].index(predict) + 65) if predict else None
                sample["extracted_answer"] = predict
            
            discipline = sample.get("discipline", "unknown")
            field = sample.get("field", "unknown")
            subfield = sample.get("subfield", "unknown")
            difficulty = sample.get("difficulty", "unknown")
            
            for level, key in [
                ('discipline', discipline),
                ('field', f"{discipline}/{field}"),
                ('subfield', f"{discipline}/{field}/{subfield}")
            ]:
                if key not in stats[level]:
                    stats[level][key] = {
                        "correct": 0, 
                        "total": 0, 
                        "miss": 0, 
                        "error": 0,
                        "discipline": discipline,
                        "field": field,
                        "subfield": subfield,
                        "difficulty": {
                            "easy": {"correct": 0, "total": 0},
                            "middle": {"correct": 0, "total": 0},
                            "hard": {"correct": 0, "total": 0}
                        }
                    }
                
                stats[level][key]["total"] += 1
                stats[level][key]["difficulty"][difficulty]["total"] += 1
                
                answer_letter = sample["answer_letter"]
                
                if predict and answer_letter == predict:
                    acc += 1
                    acc_difficulty[difficulty] += 1
                    sample["status"] = "correct"
                    stats[level][key]["correct"] += 1
                    stats[level][key]["difficulty"][difficulty]["correct"] += 1
                elif predict == None or predict == "":
                    miss += 1
                    sample["status"] = "miss"
                    stats[level][key]["miss"] += 1
                elif predict == 'error':
                    err += 1
                    sample["status"] = "error"
                    stats[level][key]["error"] += 1
                else:
                    sample["status"] = "incorrect"
                count += 1
                count_difficulty[difficulty] += 1
        
        regex_time = time.time() - regex_start_time
        
        return {
            'file_name': file_name,
            'model_name': model_name,
            'split': split,
            'mode': mode,
            'data': data,
            'read_time': read_time,
            'regex_time': regex_time,
            'sample_count': len(data),
            'results': {
                'accuracy': acc / count if count > 0 else 0,
                'error_rate': err / count if count > 0 else 0,
                'miss_rate': miss / count if count > 0 else 0,
                'hard_accuracy': acc_difficulty["hard"] / count_difficulty["hard"] if count_difficulty["hard"] > 0 else 0,
                'middle_accuracy': acc_difficulty["middle"] / count_difficulty["middle"] if count_difficulty["middle"] > 0 else 0,
                'easy_accuracy': acc_difficulty["easy"] / count_difficulty["easy"] if count_difficulty["easy"] > 0 else 0
            },
            'stats': stats
        }
    except Exception as e:
        print(f"Error processing file {file_name}: {str(e)}")
        return None

def main(args):
    results_table = PrettyTable()
    results_table.field_names = ["Model", "Split", "Mode", "Accuracy", "Errors", "Miss"]
    
    model_results = {}
    final_results = {}
    hierarchy_stats = {
        'discipline': {},
        'field': {},
        'subfield': {}
    }
    
    if args.evaluate_all:
        files = sorted([f for f in os.listdir(args.output_dir) if f.endswith('.jsonl')])
        output_suffix = args.split + '_all_models'
    else:
        if not isinstance(args.model_name, list):
            args.model_name = [args.model_name]
        
        files = []
        for model in args.model_name:
            for mode in args.mode:
                file_name = f"{model}_{args.split}_{mode}.jsonl"
                if os.path.exists(os.path.join(args.output_dir, file_name)):
                    files.append(file_name)
        output_suffix = args.split + '_' + '_'.join(args.model_name)

    max_workers = min(multiprocessing.cpu_count(), len(files))
    print(f"Using ProcessPool with {max_workers} workers")
    
    results_list = []
    total_samples = 0
    total_time = 0
    
    with multiprocessing.Pool(processes=max_workers) as pool:
        process_func = partial(process_single_file, args=args)
        for result in tqdm(
            pool.imap_unordered(process_func, files),
            total=len(files),
            desc="Processing files"
        ):
            if result is not None:
                results_list.append(result)
                total_samples += result['sample_count']
                total_time += result['read_time'] + result['regex_time']
                avg_speed = total_samples / total_time if total_time > 0 else 0
                
                tqdm.write(
                    f"File {result['file_name']} completed - "
                    f"Samples: {result['sample_count']} "
                    f"(avg {avg_speed:.1f} samples/sec)"
                )

    for result in results_list:
        model_name = result['model_name']
        mode = result['mode']
        stats = result['stats']
        
        results_table.add_row([
            model_name,
            result['split'],
            mode,
            f"{result['results']['accuracy']:.2%}",
            f"{result['results']['error_rate']:.2%}",
            f"{result['results']['miss_rate']:.2%}"
        ])
        
        if model_name not in model_results:
            model_results[model_name] = {}
        model_results[model_name][mode] = stats
        
        for level in ['discipline', 'field', 'subfield']:
            for key, data in stats[level].items():
                if key not in hierarchy_stats[level]:
                    hierarchy_stats[level][key] = data.copy()
                    hierarchy_stats[level][key]['model_stats'] = {}
                
                if model_name not in hierarchy_stats[level][key]['model_stats']:
                    hierarchy_stats[level][key]['model_stats'][model_name] = {}
                hierarchy_stats[level][key]['model_stats'][model_name][mode] = data
        
        if model_name not in final_results:
            final_results[model_name] = {}
        
        if mode not in final_results[model_name]:
            final_results[model_name][mode] = {
                "accuracy": result['results']['accuracy'],
                "errors": result['results']['error_rate'],
                "miss": result['results']['miss_rate'],
                "accuracy_hard": result['results']['hard_accuracy'],
                "accuracy_middle": result['results']['middle_accuracy'],
                "accuracy_easy": result['results']['easy_accuracy'],
                "categories": {}
            }
        
        # 更新categories
        categories_dict = final_results[model_name][mode]["categories"]
        for hierarchy_level, level_stats in stats.items():
            if hierarchy_level == "discipline" or hierarchy_level == "field":
                continue
            
            for field, field_stats in level_stats.items():
                if field not in categories_dict:
                    categories_dict[field] = {
                        "correct": field_stats["correct"],
                        "total": field_stats["total"],
                        "error": field_stats["error"],
                        "miss": field_stats["miss"],
                        "correct_hard": field_stats["difficulty"]["hard"]["correct"],
                        "total_hard": field_stats["difficulty"]["hard"]["total"],
                        "correct_middle": field_stats["difficulty"]["middle"]["correct"],
                        "total_middle": field_stats["difficulty"]["middle"]["total"],
                        "correct_easy": field_stats["difficulty"]["easy"]["correct"],
                        "total_easy": field_stats["difficulty"]["easy"]["total"]
                    }

    # Sort and print results
    results_rows = sorted(results_table._rows, key=lambda x: x[0])
    results_table.clear_rows()
    for row in results_rows:
        results_table.add_row(row)
    print(results_table)

    os.makedirs(args.save_dir, exist_ok=True)
    
    if args.excel_output:
        output_file = os.path.join(
            args.save_dir, 
            f'results_{output_suffix}.xlsx'
        )
        create_excel_report_from_stats(
            model_results,
            hierarchy_stats,
            output_file
        )
    
    if args.json_output:
        json_output_file = os.path.join(
            args.save_dir, 
            f'results_{output_suffix}.json'
        )
        with open(json_output_file, 'w', encoding='utf-8') as f:
            json.dump(final_results, f, ensure_ascii=False, indent=2)
        print(f"JSON results saved to: {json_output_file}")

    for raw_result in tqdm(results_list, desc="Saving processed results", leave=True):
        # Save processed data with status
        os.makedirs(args.save_dir, exist_ok=True)
        save_path = os.path.join(args.save_dir, raw_result['file_name'])
        with open(save_path, "w") as file:
            for sample in raw_result['data']:
                json.dump(sample, file)
                file.write("\n")

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Calculate accuracy for different modes and splits.")
    parser.add_argument('--model_name', type=str, nargs='+', default=[], help='Model names to use')
    parser.add_argument('--split', type=str, default='SuperGPQA-all', help='Data split to use')
    parser.add_argument('--mode', nargs='+', default=['zero-shot', 'five-shot'], help='Modes to use for data loading')
    parser.add_argument('--output_dir', type=str, default='results/gpqa', help='Directory to read result files from')
    parser.add_argument('--save_dir', type=str, default='results_with_status/gpqa', help='Directory to save result files with status')
    parser.add_argument('--evaluate_all', action='store_true', help='Evaluate all files in the output directory')
    parser.add_argument('--excel_output', action='store_true', help='Generate Excel report with field_final-wise results')
    parser.add_argument('--json_output', action='store_true', help='Generate JSON file with detailed results')
    
    args = parser.parse_args()
    
    # Validate parameters
    if not args.evaluate_all and not args.model_name:
        parser.error("Either --evaluate_all or --model_name must be specified")
    
    main(args)